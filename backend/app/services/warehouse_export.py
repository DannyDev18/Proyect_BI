# backend/app/services/warehouse_export.py
"""Export a Excel de los reportes tipados del módulo Bodega (Fase 5,
docs/features/plan_actualizacion_modulo_bodega.md §5.2). Consume el contrato
`schemas.warehouse.ReporteBodegaResponse` (resumen ejecutivo + secciones con columnas
de negocio ya definidas) -- ya no es un volcado genérico de JSON: hoja "Resumen" con
KPIs interpretados + filtros aplicados, y una hoja por sección con encabezados en
español, formato de moneda/porcentaje, autofiltro y resaltado de prioridad alta."""
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_RESALTAR_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_RESALTAR_VALORES = {"Alta", "Crítico"}

# Nombre de hoja de Excel: máx 31 caracteres y sin []:*?/\
_INVALIDOS_HOJA = set('[]:*?/\\')


def _nombre_hoja(nombre: str) -> str:
    limpio = "".join(c for c in nombre if c not in _INVALIDOS_HOJA)
    return (limpio.strip() or "Hoja")[:31]


def _formatear_celda(valor: Any, tipo: str) -> Any:
    if valor is None:
        return "—"
    if tipo == "moneda" and isinstance(valor, (int, float)):
        return f"${valor:,.2f}"
    if tipo == "porcentaje" and isinstance(valor, (int, float)):
        return f"{valor:+.1f}%"
    return valor


def _escribir_seccion(ws: Worksheet, seccion: dict[str, Any]) -> None:
    columnas = seccion["columnas"]
    filas = seccion["filas"]
    ws.append([col["etiqueta"] for col in columnas])
    for celda in ws[1]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
        celda.alignment = Alignment(horizontal="center")

    resaltar_key = seccion.get("resaltar_key")
    if not filas:
        ws.append(["(sin datos con los filtros actuales)"])
    for fila in filas:
        ws.append([_formatear_celda(fila.get(col["key"]), col["tipo"]) for col in columnas])
        if resaltar_key and str(fila.get(resaltar_key)) in _RESALTAR_VALORES:
            for celda in ws[ws.max_row]:
                celda.fill = _RESALTAR_FILL

    if filas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{ws.max_row}"
    for idx, col in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(40, len(col["etiqueta"]) + 6))
    ws.freeze_panes = "A2"


def reporte_a_excel(reporte: dict[str, Any]) -> bytes:
    """Convierte el contrato tipado de un reporte de bodega en un XLSX en memoria:
    hoja "Resumen" (KPIs interpretados + filtros aplicados + fecha) + una hoja por
    sección con formato de negocio y autofiltro."""
    wb = Workbook()
    resumen = wb.active
    resumen.title = "Resumen"
    resumen.append([reporte["titulo"]])
    resumen["A1"].font = Font(bold=True, size=14)
    resumen.append([f"Generado: {reporte['generado_en']}"])
    resumen.append([reporte["interpretacion"]])
    resumen["A3"].alignment = Alignment(wrap_text=True)
    resumen.append([])

    resumen.append(["Indicador", "Valor"])
    for celda in resumen[resumen.max_row]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
    for kpi in reporte["resumen_ejecutivo"]:
        resumen.append([kpi["etiqueta"], kpi["valor"]])
    resumen.append([])

    filtros = reporte.get("filtros_aplicados") or {}
    resumen.append(["Filtros aplicados", "Valor"])
    for celda in resumen[resumen.max_row]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
    if filtros:
        for clave, valor in filtros.items():
            resumen.append([clave.replace("_", " ").title(), valor])
    else:
        resumen.append(["(ninguno — todas las bodegas/categorías/proveedores)", ""])

    resumen.column_dimensions["A"].width = 45
    resumen.column_dimensions["B"].width = 40

    for seccion in reporte["secciones"]:
        _escribir_seccion(wb.create_sheet(_nombre_hoja(seccion["titulo"])), seccion)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
